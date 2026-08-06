from openpyxl import load_workbook
import os
from openpyxl.styles import Alignment,  PatternFill

def atualizar_dia(tempo,periodo, linha_busca, situacao,wb,aba,nome_arquivo):
    
    prog = 1

    while True:
        if aba.cell(row=linha_busca-1, column=prog).value != None:
            prog += 1
        else:
            break   
        
    qtd=aba.cell(row=linha_busca, column=1, value=prog)
    qtd.alignment= Alignment(horizontal="center", vertical="center")
        
    sit=aba.cell(row=linha_busca-1, column=prog, value=situacao)
    sit.alignment= Alignment(horizontal="center", vertical="center")

    tem=aba.cell(row=linha_busca-2, column=prog, value=int(tempo))
    tem.alignment= Alignment(horizontal="center", vertical="center")

    per=aba.cell(row=linha_busca-3, column=prog, value=periodo)
    per.alignment= Alignment(horizontal="center", vertical="center")
    
    if periodo == "M":
        per.fill= PatternFill(start_color="EEFB6F", end_color="EEFB6F", fill_type="solid")
    elif periodo == "T":
        per.fill= PatternFill(start_color="FFC16B", end_color="FFC16B", fill_type="solid")
    else:
        per.fill= PatternFill(start_color="7B6DFD", end_color="7B6DFD", fill_type="solid")
    
    wb.save(nome_arquivo)
    
def novo_dia(tempo,periodo, linha_busca, situacao,wb,aba,Dia_hoje,nome_arquivo):

    dia=aba.cell(row=linha_busca+1, column=1, value=Dia_hoje)
    dia.alignment= Alignment(horizontal="center", vertical="center")
    dia.fill= PatternFill(start_color="B5EDBA", end_color="B5EDBA", fill_type="solid")
    per=aba.cell(row=linha_busca+2, column=1, value=periodo)
    per.alignment=Alignment(horizontal="center", vertical="center")

    if periodo == "M":
        per.fill= PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    elif periodo == "T":
        per.fill= PatternFill(start_color="FFC16B", end_color="FFC16B", fill_type="solid")
    else:
        per.fill= PatternFill(start_color="7B6DFD", end_color="7B6DFD", fill_type="solid")

    tem=aba.cell(row=linha_busca+3, column=1, value=int(tempo))
    tem.alignment= Alignment(horizontal="center", vertical="center")

    sit=aba.cell(row=linha_busca+4, column=1, value=situacao)
    sit.alignment= Alignment(horizontal="center", vertical="center")
    aba.cell(row=linha_busca+5, column=1, value=1)

    wb.save(nome_arquivo)







    