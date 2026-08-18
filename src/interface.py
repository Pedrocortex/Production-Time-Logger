import tkinter as tk
from tkinter import messagebox
from guuar import atualizar_dia, novo_dia
import os
from openpyxl import load_workbook
from datetime import datetime

switcher = False
Dia_hoje = datetime.now().strftime("%d/%m/%Y")
def atualiza():

    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    nome_arquivo = os.path.join(diretorio_atual, "Dados1.xlsx")
    wb = load_workbook(nome_arquivo)
    aba = wb.active

    return wb,aba,nome_arquivo, 


def enviar_dados(periodo,tempo, situacao,janela):    
    wb,aba,nome_arquivo = atualiza()
    
    ultima_linha = aba["M1"].value
    data_celula = aba.cell(row=ultima_linha-4, column=1).value
    
    if not tempo:
        messagebox.showwarning("Aviso", "Por favor, digite o Tempo!")
        return
    if Dia_hoje == str(data_celula):
        atualizar_dia(tempo, periodo, ultima_linha,situacao,wb,aba,nome_arquivo)
    else:
        
        novo_dia(tempo,periodo, ultima_linha,situacao,wb,aba,Dia_hoje,nome_arquivo)

    janela.destroy()
    switcher=True
    recebe_dados(periodo,situacao )

    return switcher


# 1. Cria a janela principal
def recebe_dados(p,s):
        
    janela = tk.Tk()
    janela.title("Portal de Cadastro")
    janela.geometry("300x410") # Define o tamanho da janela (Largura x Altura)
    
    btn_atualizar = tk.Button(
        janela,
        text="Apagar número",
        command=lambda: (janela.destroy(), recebe_dados(p, s))

)

    btn_atualizar.pack(pady=5)
    # 2. Campo: Tempo
    lbl_tempo = tk.Label(janela, text="Tempo de operação (Minutos):")
    lbl_tempo.pack(pady=5)
    
    entrada_tempo = tk.Entry(janela)
    entrada_tempo.pack(pady=5)

    # 3. Campo: Turno (Usando botões de rádio para M, T, N)
    lbl_periodo = tk.Label(janela, text="Selecione o Turno:")
    lbl_periodo.pack(pady=5)

    var_periodo = tk.StringVar(value=p) # Define "M" como padrão inicial

    tk.Radiobutton(janela, text="Matutino (M)", variable=var_periodo, value="M").pack()
    tk.Radiobutton(janela, text="Vespertino (T)", variable=var_periodo, value="T").pack()
    tk.Radiobutton(janela, text="Noturno (N)", variable=var_periodo, value="N").pack()

    # 4. Campo: Rítimo de produção (Usando botões de rádio para M, T, N)
    lbl_situacao = tk.Label(janela, text="Selecione o rítimo de produção:")
    lbl_situacao.pack(pady=5)

    var_situacao = tk.StringVar(value=s) 

    tk.Radiobutton(janela, text="Normal", variable=var_situacao  , value="Normal").pack()
    tk.Radiobutton(janela, text="Intenso", variable=var_situacao, value="Intenso").pack()
    tk.Radiobutton(janela, text="Crítico", variable=var_situacao, value="Crítico").pack()
    # 4. Botão de Enviar
    btn_enviar = tk.Button(
        janela,
        text="Gravar Dados",
        command=lambda: enviar_dados(
        var_periodo.get(),
        entrada_tempo.get(),
        var_situacao.get(),
        janela
        ),
        bg="#4A90E2",
        fg="white"
    )

    btn_enviar.pack(pady=20)

    janela.mainloop()
    
if not switcher:
    recebe_dados ("M","Normal")
