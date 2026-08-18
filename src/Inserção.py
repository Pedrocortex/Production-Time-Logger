from openpyxl import load_workbook
import os
from openpyxl.styles import Alignment,  PatternFill

def atualizar_dia(tempo,periodo, ultima_linha, situacao,wb,aba,nome_arquivo):
    
    prog = aba.cell(row=ultima_linha, column=1).value+1  
    
    qtd=aba.cell(row=ultima_linha, column=1, value=prog)
    qtd.alignment= Alignment(horizontal="center", vertical="center")

    sit=aba.cell(row=ultima_linha-1, column=prog, value=situacao)
    sit.alignment= Alignment(horizontal="center", vertical="center")

    tem=aba.cell(row=ultima_linha-2, column=prog, value=int(tempo))
    tem.alignment= Alignment(horizontal="center", vertical="center")

    per=aba.cell(row=ultima_linha-3, column=prog, value=periodo)
    per.alignment= Alignment(horizontal="center", vertical="center")
    
    if periodo == "M":
        per.fill= PatternFill(start_color="EEFB6F", end_color="EEFB6F", fill_type="solid")
    elif periodo == "T":
        per.fill= PatternFill(start_color="FFC16B", end_color="FFC16B", fill_type="solid")
    else:
        per.fill= PatternFill(start_color="7B6DFD", end_color="7B6DFD", fill_type="solid")
    
    wb.save(nome_arquivo)
    
def novo_dia(tempo,periodo, ultima_linha, situacao,wb,aba,Dia_hoje,nome_arquivo):
     if aba.cell(row=ultima_linha, column=1).value is None:
            ultima_linha = ultima_linha - 5
         
    dia=aba.cell(row=ultima_linha+1, column=1, value=Dia_hoje)
    dia.alignment= Alignment(horizontal="center", vertical="center")
    dia.fill= PatternFill(start_color="B5EDBA", end_color="B5EDBA", fill_type="solid")

    per=aba.cell(row=ultima_linha+2, column=1, value=periodo)
    per.alignment=Alignment(horizontal="center", vertical="center")

    if periodo == "M":
        per.fill= PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    elif periodo == "T":
        per.fill= PatternFill(start_color="FFC16B", end_color="FFC16B", fill_type="solid")
    else:
        per.fill= PatternFill(start_color="7B6DFD", end_color="7B6DFD", fill_type="solid")

    tem=aba.cell(row=ultima_linha+3, column=1, value=int(tempo))
    tem.alignment= Alignment(horizontal="center", vertical="center")

    sit=aba.cell(row=ultima_linha+4, column=1, value=situacao)
    sit.alignment= Alignment(horizontal="center", vertical="center")
    aba.cell(row=ultima_linha+5, column=1, value=1)
    aba["M1"] = ultima_linha+5
    wb.save(nome_arquivo)







    
